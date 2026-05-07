# routes/rendiconto.py
# Blueprint Rendiconto Studio — Volume 4
# Aggrega pagamenti, movimenti studio, giroconti per anno/mese.

from flask import Blueprint, jsonify, request, session, send_file
from database import get_db
from functools import wraps
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

bp = Blueprint('rendiconto', __name__)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Non autenticato'}), 401
        return f(*args, **kwargs)
    return decorated


def _gestione_attiva(ditta: dict, anno: int, tipo: str) -> bool:
    """True se la gestione (paghe/contabilita) era attiva durante l'anno.
    Se fine_X è impostata, l'anno di fine è già considerato chiuso (>=)."""
    inizio = ditta.get(f'inizio_{tipo}')
    fine   = ditta.get(f'fine_{tipo}')
    if not inizio:
        return False
    try:
        if anno < int(str(inizio)[:4]):
            return False
        if fine and anno >= int(str(fine)[:4]):
            return False
    except Exception:
        pass
    return True


def _calcola_residuo(db, ditta_id: int, anno: str) -> float:
    """Residuo cumulativo di un cliente calcolato su tutto lo storico."""
    row = db.execute('SELECT residuo_iniziale FROM ditte WHERE id=?', (ditta_id,)).fetchone()
    res_ini = row['residuo_iniziale'] if row else 0.0

    dovuto  = db.execute('SELECT COALESCE(SUM(importo),0) FROM pratiche WHERE ditta_id=?', (ditta_id,)).fetchone()[0]
    pagato  = db.execute('SELECT COALESCE(SUM(importo),0) FROM pagamenti WHERE ditta_id=?', (ditta_id,)).fetchone()[0]
    abbuoni = db.execute("SELECT COALESCE(SUM(importo),0) FROM arrotondamenti WHERE ditta_id=? AND tipo='abbuono'", (ditta_id,)).fetchone()[0]
    addebiti= db.execute("SELECT COALESCE(SUM(importo),0) FROM arrotondamenti WHERE ditta_id=? AND tipo='addebito'", (ditta_id,)).fetchone()[0]

    return round(res_ini + dovuto - pagato - abbuoni + addebiti, 2)


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/saldi
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/saldi')
@login_required
def get_saldi():
    db = get_db()
    saldi = []

    cassa_in      = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='entrata'").fetchone()[0]
    cassa_out     = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='uscita'").fetchone()[0]
    cassa_giro_in = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='giroconto' AND giroconto_dir='entrata'").fetchone()[0]
    cassa_giro_out= db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='giroconto' AND giroconto_dir='uscita'").fetchone()[0]

    cassa_saldo_iniziale = 0.0
    try:
        row = db.execute("SELECT valore FROM impostazioni WHERE chiave='saldo_iniziale_cassa'").fetchone()
        if row:
            cassa_saldo_iniziale = float(row['valore'])
    except Exception:
        pass

    saldi.append({'id': 'cassa', 'nome': 'Cassa',
                  'saldo': round(cassa_saldo_iniziale + cassa_in - cassa_out + cassa_giro_in - cassa_giro_out, 2)})

    for banca in db.execute('SELECT * FROM banche_studio ORDER BY ordine, nome').fetchall():
        tid = f"banca_{banca['id']}"
        b_in      = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='entrata'", (tid,)).fetchone()[0]
        b_out     = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='uscita'", (tid,)).fetchone()[0]
        b_giro_in = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='giroconto' AND giroconto_dir='entrata'", (tid,)).fetchone()[0]
        b_giro_out= db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='giroconto' AND giroconto_dir='uscita'", (tid,)).fetchone()[0]
        saldi.append({'id': tid, 'nome': banca['nome'],
                      'saldo': round(banca['saldo_iniziale'] + b_in - b_out + b_giro_in - b_giro_out, 2),
                      'colore': banca['colore'] or '#6366f1'})

    db.close()
    return jsonify(saldi)


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/riepilogo?anno=XXXX
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/riepilogo')
@login_required
def get_riepilogo():
    anno = request.args.get('anno', '')
    db   = get_db()

    if anno:
        pag          = db.execute("SELECT COALESCE(SUM(importo),0) FROM pagamenti WHERE anno=?", (anno,)).fetchone()[0]
        altre_entrate= db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipo='entrata' AND strftime('%Y',data)=?", (anno,)).fetchone()[0]
        uscite       = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipo='uscita'  AND strftime('%Y',data)=?", (anno,)).fetchone()[0]
    else:
        pag          = db.execute("SELECT COALESCE(SUM(importo),0) FROM pagamenti").fetchone()[0]
        altre_entrate= db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipo='entrata'").fetchone()[0]
        uscite       = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipo='uscita'").fetchone()[0]

    tot_e = round(pag + altre_entrate, 2)
    tot_u = round(uscite, 2)
    db.close()
    return jsonify({'tot_entrate': tot_e, 'tot_uscite': tot_u, 'differenza': round(tot_e - tot_u, 2)})


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/entrate?anno=XXXX&mostra_archiviati=0
# Sezioni clienti (per categoria) + macrogruppi liberi, per mese
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/entrate')
@login_required
def get_entrate():
    anno             = request.args.get('anno', str(date.today().year))
    mostra_archiviati= request.args.get('mostra_archiviati', '0') == '1'
    db               = get_db()
    anno_str         = str(anno)

    # ── Pagamenti mensili per cliente ────────────────────────────
    pag_rows = db.execute(
        '''SELECT ditta_id, CAST(strftime('%m', data) AS INTEGER) AS mese, SUM(importo) AS tot
           FROM pagamenti WHERE anno=?
           GROUP BY ditta_id, mese''', (anno_str,)
    ).fetchall()
    pag_by_ditta = {}
    for r in pag_rows:
        pag_by_ditta.setdefault(r['ditta_id'], {})[r['mese']] = r['tot']

    # ── Clienti ──────────────────────────────────────────────────
    arch_sql = '' if mostra_archiviati else 'AND (d.archiviato IS NULL OR d.archiviato=0)'
    ditte = db.execute(
        f'''SELECT id, ragione_sociale, residuo_iniziale,
                   inizio_paghe, fine_paghe, inizio_contabilita, fine_contabilita
            FROM ditte d WHERE 1=1 {arch_sql}
            ORDER BY ragione_sociale'''
    ).fetchall()

    sezioni = {
        'paghe':      {'nome': 'CLIENTI PAGHE',               'colore': 'paghe',      'clienti': []},
        'cont':       {'nome': 'CLIENTI CONTABILITÀ',          'colore': 'cont',       'clienti': []},
        'paghe_cont': {'nome': 'CLIENTI PAGHE + CONTABILITÀ',  'colore': 'paghe_cont', 'clienti': []},
        'altro':      {'nome': 'CLIENTI (ALTRO)',               'colore': 'altro',      'clienti': []},
    }

    for d in ditte:
        d = dict(d)
        mesi_map = pag_by_ditta.get(d['id'], {})
        mesi     = [round(mesi_map.get(m, 0.0), 2) for m in range(1, 13)]
        tot_pag  = round(sum(mesi), 2)
        residuo  = _calcola_residuo(db, d['id'], anno_str)

        if tot_pag == 0 and residuo == 0:
            continue

        ha_p = _gestione_attiva(d, int(anno_str), 'paghe')
        ha_c = _gestione_attiva(d, int(anno_str), 'contabilita')

        entry = {'id': d['id'], 'nome': d['ragione_sociale'],
                 'mesi': mesi, 'tot_pagato': tot_pag, 'residuo': residuo}

        if ha_p and ha_c:
            sezioni['paghe_cont']['clienti'].append(entry)
        elif ha_p:
            sezioni['paghe']['clienti'].append(entry)
        elif ha_c:
            sezioni['cont']['clienti'].append(entry)
        else:
            sezioni['altro']['clienti'].append(entry)

    result_sezioni = []
    for key, sez in sezioni.items():
        if not sez['clienti']:
            continue
        sub_mesi = [round(sum(c['mesi'][i] for c in sez['clienti']), 2) for i in range(12)]
        result_sezioni.append({
            'id': key, 'nome': sez['nome'], 'colore': sez['colore'],
            'clienti': sez['clienti'],
            'subtotali_mesi': sub_mesi,
            'subtotale': round(sum(sub_mesi), 2),
            'subtotale_residui': round(sum(c['residuo'] for c in sez['clienti']), 2),
        })

    # ── Macrogruppi entrate liberi (movimenti_studio, non-clienti) ─
    ms_rows = db.execute(
        '''SELECT macrogruppo_id, macrogruppo_nome, sottovoce_id, sottovoce_nome,
                  CAST(strftime('%m', data) AS INTEGER) AS mese, SUM(importo) AS tot
           FROM movimenti_studio
           WHERE tipo='entrata' AND strftime('%Y', data)=?
             AND (macrogruppo_id IS NULL OR macrogruppo_id != 'clienti')
           GROUP BY macrogruppo_id, sottovoce_id, mese
           ORDER BY macrogruppo_nome, sottovoce_nome, mese''', (anno_str,)
    ).fetchall()

    mg_map = {}
    for r in ms_rows:
        mgid   = r['macrogruppo_id'] or '__altro__'
        mgnome = r['macrogruppo_nome'] or 'Altre entrate'
        svid   = r['sottovoce_id']   or '__sv__'
        svnome = r['sottovoce_nome'] or '—'
        mg_map.setdefault(mgid, {'id': mgid, 'nome': mgnome, 'sottovoci': {}})
        mg_map[mgid]['sottovoci'].setdefault(svid, {'id': svid, 'nome': svnome, 'mesi': [0.0]*12})
        mg_map[mgid]['sottovoci'][svid]['mesi'][r['mese'] - 1] = round(r['tot'], 2)

    result_mg = []
    for mg in mg_map.values():
        sottovoci = list(mg['sottovoci'].values())
        for sv in sottovoci:
            sv['totale'] = round(sum(sv['mesi']), 2)
        sub_mesi = [round(sum(sv['mesi'][i] for sv in sottovoci), 2) for i in range(12)]
        result_mg.append({'id': mg['id'], 'nome': mg['nome'],
                          'sottovoci': sottovoci,
                          'subtotali_mesi': sub_mesi,
                          'subtotale': round(sum(sub_mesi), 2)})

    # ── Totali generali ──────────────────────────────────────────
    totali_mesi = [0.0] * 12
    for s in result_sezioni:
        for i in range(12): totali_mesi[i] += s['subtotali_mesi'][i]
    for m in result_mg:
        for i in range(12): totali_mesi[i] += m['subtotali_mesi'][i]
    totali_mesi = [round(v, 2) for v in totali_mesi]

    db.close()
    return jsonify({
        'sezioni_clienti': result_sezioni,
        'macrogruppi':     result_mg,
        'totali_mesi':     totali_mesi,
        'totale_annuale':  round(sum(totali_mesi), 2),
        'totale_residui':  round(sum(s.get('subtotale_residui', 0) for s in result_sezioni), 2),
    })


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/uscite?anno=XXXX
# Gerarchia macrogruppo → sottovoce, distribuzione mensile
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/uscite')
@login_required
def get_uscite():
    anno     = request.args.get('anno', str(date.today().year))
    anno_str = str(anno)
    db       = get_db()

    rows = db.execute(
        '''SELECT macrogruppo_id, macrogruppo_nome, sottovoce_id, sottovoce_nome,
                  CAST(strftime('%m', data) AS INTEGER) AS mese, SUM(importo) AS tot
           FROM movimenti_studio
           WHERE tipo='uscita' AND strftime('%Y', data)=?
           GROUP BY macrogruppo_id, sottovoce_id, mese
           ORDER BY macrogruppo_nome, sottovoce_nome, mese''', (anno_str,)
    ).fetchall()

    mg_map = {}
    for r in rows:
        mgid   = r['macrogruppo_id']   or '__altro__'
        mgnome = r['macrogruppo_nome'] or 'Altre uscite'
        svid   = r['sottovoce_id']     or '__sv__'
        svnome = r['sottovoce_nome']   or '—'
        mg_map.setdefault(mgid, {'id': mgid, 'nome': mgnome, 'sottovoci': {}})
        mg_map[mgid]['sottovoci'].setdefault(svid, {'id': svid, 'nome': svnome, 'mesi': [0.0]*12})
        mg_map[mgid]['sottovoci'][svid]['mesi'][r['mese'] - 1] = round(r['tot'], 2)

    macrogruppi = []
    totali_mesi = [0.0] * 12
    for mg in mg_map.values():
        sottovoci = list(mg['sottovoci'].values())
        for sv in sottovoci:
            sv['totale'] = round(sum(sv['mesi']), 2)
        sub_mesi = [round(sum(sv['mesi'][i] for sv in sottovoci), 2) for i in range(12)]
        for i in range(12): totali_mesi[i] += sub_mesi[i]
        macrogruppi.append({'id': mg['id'], 'nome': mg['nome'],
                            'sottovoci': sottovoci,
                            'subtotali_mesi': sub_mesi,
                            'subtotale': round(sum(sub_mesi), 2)})

    totali_mesi = [round(v, 2) for v in totali_mesi]
    db.close()
    return jsonify({
        'macrogruppi':    macrogruppi,
        'totali_mesi':    totali_mesi,
        'totale_annuale': round(sum(totali_mesi), 2),
    })


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/giroconti?anno=XXXX
# Solo la "gamba uscita" di ogni giroconto.
# Da = conto origine (tipologia), A = conto destinazione (sottovoce_nome)
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/giroconti')
@login_required
def get_giroconti():
    anno     = request.args.get('anno', str(date.today().year))
    anno_str = str(anno)
    db       = get_db()

    rows = db.execute(
        '''SELECT m.data,
                  CASE WHEN m.tipologia='cassa' THEN 'Cassa'
                       ELSE COALESCE(b.nome, m.tipologia) END AS da,
                  m.sottovoce_nome  AS a,
                  m.descrizione,
                  m.importo
           FROM movimenti_studio m
           LEFT JOIN banche_studio b ON ('banca_' || b.id = m.tipologia)
           WHERE m.tipo='giroconto' AND m.giroconto_dir='uscita'
             AND strftime('%Y', m.data)=?
           ORDER BY m.data DESC, m.id DESC''', (anno_str,)
    ).fetchall()

    result = [{'data': r['data'], 'da': r['da'] or '—',
               'a': r['a'] or '—', 'descrizione': r['descrizione'] or '',
               'importo': round(r['importo'], 2)} for r in rows]
    db.close()
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/export-pdf?anno=XXXX
# Genera Rendiconto_<anno>.pdf  A4 landscape, 10mm margini
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/export-pdf')
@login_required
def export_pdf():
    anno     = request.args.get('anno', str(date.today().year))
    anno_str = str(anno)
    db       = get_db()

    # ── raccolta dati ────────────────────────────────────────────
    # saldi
    saldi_list = _query_saldi(db)
    # riepilogo
    riepilogo  = _query_riepilogo(db, anno_str)
    # entrate
    entrate    = _query_entrate(db, anno_str, mostra_archiviati=True)
    # uscite
    uscite     = _query_uscite(db, anno_str)
    # giroconti
    giroconti  = _query_giroconti(db, anno_str)
    db.close()

    buf = _genera_pdf(anno_str, saldi_list, riepilogo, entrate, uscite, giroconti)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'Rendiconto_{anno_str}.pdf')


# ── helper query (riutilizzabili) ────────────────────────────────
def _query_saldi(db):
    saldi = []
    ci  = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='entrata'").fetchone()[0]
    co  = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='uscita'").fetchone()[0]
    cgi = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='giroconto' AND giroconto_dir='entrata'").fetchone()[0]
    cgo = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia='cassa' AND tipo='giroconto' AND giroconto_dir='uscita'").fetchone()[0]
    ini = 0.0
    try:
        r = db.execute("SELECT valore FROM impostazioni WHERE chiave='saldo_iniziale_cassa'").fetchone()
        if r: ini = float(r['valore'])
    except Exception: pass
    saldi.append({'nome': 'Cassa', 'saldo': round(ini + ci - co + cgi - cgo, 2)})
    for b in db.execute('SELECT * FROM banche_studio ORDER BY ordine, nome').fetchall():
        tid = f"banca_{b['id']}"
        bi  = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='entrata'", (tid,)).fetchone()[0]
        bo  = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='uscita'", (tid,)).fetchone()[0]
        bgi = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='giroconto' AND giroconto_dir='entrata'", (tid,)).fetchone()[0]
        bgo = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipologia=? AND tipo='giroconto' AND giroconto_dir='uscita'", (tid,)).fetchone()[0]
        saldi.append({'nome': b['nome'], 'saldo': round(b['saldo_iniziale'] + bi - bo + bgi - bgo, 2)})
    return saldi


def _query_riepilogo(db, anno_str):
    pag   = db.execute("SELECT COALESCE(SUM(importo),0) FROM pagamenti WHERE anno=?", (anno_str,)).fetchone()[0]
    altre = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipo='entrata' AND strftime('%Y',data)=?", (anno_str,)).fetchone()[0]
    usc   = db.execute("SELECT COALESCE(SUM(importo),0) FROM movimenti_studio WHERE tipo='uscita'  AND strftime('%Y',data)=?", (anno_str,)).fetchone()[0]
    tot_e = round(pag + altre, 2)
    tot_u = round(usc, 2)
    return {'tot_entrate': tot_e, 'tot_uscite': tot_u, 'differenza': round(tot_e - tot_u, 2)}


def _query_entrate(db, anno_str, mostra_archiviati=True):
    arch  = '' if mostra_archiviati else 'AND (d.archiviato IS NULL OR d.archiviato=0)'
    ditte = db.execute(f'SELECT id,ragione_sociale,residuo_iniziale,inizio_paghe,fine_paghe,inizio_contabilita,fine_contabilita FROM ditte d WHERE 1=1 {arch} ORDER BY ragione_sociale').fetchall()
    pag_r = db.execute("SELECT ditta_id, CAST(strftime('%m',data) AS INTEGER) AS mese, SUM(importo) AS tot FROM pagamenti WHERE anno=? GROUP BY ditta_id, mese", (anno_str,)).fetchall()
    pag_d = {}
    for r in pag_r: pag_d.setdefault(r['ditta_id'], {})[r['mese']] = r['tot']

    sezioni = {'paghe':{'nome':'CLIENTI PAGHE','colore':'paghe','clienti':[]},'cont':{'nome':'CLIENTI CONTABILITÀ','colore':'cont','clienti':[]},'paghe_cont':{'nome':'CLIENTI PAGHE + CONTABILITÀ','colore':'paghe_cont','clienti':[]},'altro':{'nome':'CLIENTI (ALTRO)','colore':'altro','clienti':[]}}
    for d in ditte:
        d = dict(d)
        mm = pag_d.get(d['id'], {})
        mesi = [round(mm.get(m, 0.0), 2) for m in range(1, 13)]
        tot  = round(sum(mesi), 2)
        res  = _calcola_residuo(db, d['id'], anno_str)
        if tot == 0 and res == 0: continue
        ha_p = _gestione_attiva(d, int(anno_str), 'paghe')
        ha_c = _gestione_attiva(d, int(anno_str), 'contabilita')
        entry = {'id': d['id'], 'nome': d['ragione_sociale'], 'mesi': mesi, 'tot_pagato': tot, 'residuo': res}
        if ha_p and ha_c: sezioni['paghe_cont']['clienti'].append(entry)
        elif ha_p:        sezioni['paghe']['clienti'].append(entry)
        elif ha_c:        sezioni['cont']['clienti'].append(entry)
        else:             sezioni['altro']['clienti'].append(entry)

    result_s = []
    for sez in sezioni.values():
        if not sez['clienti']: continue
        sub = [round(sum(c['mesi'][i] for c in sez['clienti']), 2) for i in range(12)]
        result_s.append({'nome': sez['nome'], 'colore': sez['colore'], 'clienti': sez['clienti'],
                         'subtotali_mesi': sub, 'subtotale': round(sum(sub), 2),
                         'subtotale_residui': round(sum(c['residuo'] for c in sez['clienti']), 2)})

    ms = db.execute("SELECT macrogruppo_id, macrogruppo_nome, sottovoce_id, sottovoce_nome, CAST(strftime('%m',data) AS INTEGER) AS mese, SUM(importo) AS tot FROM movimenti_studio WHERE tipo='entrata' AND strftime('%Y',data)=? AND (macrogruppo_id IS NULL OR macrogruppo_id!='clienti') GROUP BY macrogruppo_id,sottovoce_id,mese ORDER BY macrogruppo_nome,sottovoce_nome,mese", (anno_str,)).fetchall()
    mg_map = {}
    for r in ms:
        mgid = r['macrogruppo_id'] or '__altro__'; mgnome = r['macrogruppo_nome'] or 'Altre entrate'
        svid = r['sottovoce_id'] or '__sv__'; svnome = r['sottovoce_nome'] or '—'
        mg_map.setdefault(mgid, {'nome': mgnome, 'sottovoci': {}})
        mg_map[mgid]['sottovoci'].setdefault(svid, {'nome': svnome, 'mesi': [0.0]*12})
        mg_map[mgid]['sottovoci'][svid]['mesi'][r['mese']-1] = round(r['tot'], 2)
    result_mg = []
    for mg in mg_map.values():
        svs = list(mg['sottovoci'].values())
        for sv in svs: sv['totale'] = round(sum(sv['mesi']), 2)
        sub = [round(sum(sv['mesi'][i] for sv in svs), 2) for i in range(12)]
        result_mg.append({'nome': mg['nome'], 'sottovoci': svs, 'subtotali_mesi': sub, 'subtotale': round(sum(sub), 2)})

    tot_mesi = [round(sum(s['subtotali_mesi'][i] for s in result_s) + sum(m['subtotali_mesi'][i] for m in result_mg), 2) for i in range(12)]
    return {'sezioni_clienti': result_s, 'macrogruppi': result_mg, 'totali_mesi': tot_mesi,
            'totale_annuale': round(sum(tot_mesi), 2),
            'totale_residui': round(sum(s.get('subtotale_residui', 0) for s in result_s), 2)}


def _query_uscite(db, anno_str):
    rows = db.execute("SELECT macrogruppo_id, macrogruppo_nome, sottovoce_id, sottovoce_nome, CAST(strftime('%m',data) AS INTEGER) AS mese, SUM(importo) AS tot FROM movimenti_studio WHERE tipo='uscita' AND strftime('%Y',data)=? GROUP BY macrogruppo_id,sottovoce_id,mese ORDER BY macrogruppo_nome,sottovoce_nome,mese", (anno_str,)).fetchall()
    mg_map = {}
    for r in rows:
        mgid = r['macrogruppo_id'] or '__altro__'; mgnome = r['macrogruppo_nome'] or 'Altre uscite'
        svid = r['sottovoce_id'] or '__sv__'; svnome = r['sottovoce_nome'] or '—'
        mg_map.setdefault(mgid, {'nome': mgnome, 'sottovoci': {}})
        mg_map[mgid]['sottovoci'].setdefault(svid, {'nome': svnome, 'mesi': [0.0]*12})
        mg_map[mgid]['sottovoci'][svid]['mesi'][r['mese']-1] = round(r['tot'], 2)
    mgs = []
    tot = [0.0]*12
    for mg in mg_map.values():
        svs = list(mg['sottovoci'].values())
        for sv in svs: sv['totale'] = round(sum(sv['mesi']), 2)
        sub = [round(sum(sv['mesi'][i] for sv in svs), 2) for i in range(12)]
        for i in range(12): tot[i] += sub[i]
        mgs.append({'nome': mg['nome'], 'sottovoci': svs, 'subtotali_mesi': sub, 'subtotale': round(sum(sub), 2)})
    tot = [round(v, 2) for v in tot]
    return {'macrogruppi': mgs, 'totali_mesi': tot, 'totale_annuale': round(sum(tot), 2)}


def _query_giroconti(db, anno_str):
    rows = db.execute("SELECT m.data, CASE WHEN m.tipologia='cassa' THEN 'Cassa' ELSE COALESCE(b.nome, m.tipologia) END AS da, m.sottovoce_nome AS a, m.descrizione, m.importo FROM movimenti_studio m LEFT JOIN banche_studio b ON ('banca_'||b.id=m.tipologia) WHERE m.tipo='giroconto' AND m.giroconto_dir='uscita' AND strftime('%Y',m.data)=? ORDER BY m.data DESC, m.id DESC", (anno_str,)).fetchall()
    return [{'data': r['data'], 'da': r['da'] or '—', 'a': r['a'] or '—', 'descrizione': r['descrizione'] or '', 'importo': round(r['importo'], 2)} for r in rows]


# ── generazione PDF ──────────────────────────────────────────────
def _genera_pdf(anno_str, saldi_list, riepilogo, entrate, uscite, giroconti):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    MESI_BREVI = ['Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic']
    C_DARK   = colors.HexColor('#1a202c')
    C_GREEN  = colors.HexColor('#16a34a')
    C_RED    = colors.HexColor('#dc2626')
    C_PURPLE = colors.HexColor('#7c3aed')
    C_MUTED  = colors.HexColor('#94a3b8')
    C_PAGHE      = colors.HexColor('#dcfce7')
    C_CONT       = colors.HexColor('#dbeafe')
    C_PAGHE_CONT = colors.HexColor('#fed7aa')
    C_ALTRO      = colors.HexColor('#f1f5f9')
    C_MACRO      = colors.HexColor('#e2e8f0')
    C_HEADER_ROW = colors.HexColor('#f8fafc')
    C_TOTAL      = colors.HexColor('#1e293b')

    def _n(v, color=None):
        """Formatta numero con separatore migliaia italiano."""
        if v == 0:
            return '—'
        s = f'{abs(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return ('-' if v < 0 else '') + s

    PAGE = landscape(A4)
    M    = 10 * mm
    W    = PAGE[0] - 2 * M  # 277mm contenuto

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=M, rightMargin=M,
                            topMargin=M, bottomMargin=M)

    sNorm  = ParagraphStyle('n', fontName='Helvetica',      fontSize=8,  leading=10)
    sBold  = ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=8,  leading=10)
    sTitle = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.white, alignment=TA_CENTER)
    sSub   = ParagraphStyle('s', fontName='Helvetica',      fontSize=9,  leading=11, textColor=colors.HexColor('#64748b'))
    sRight = ParagraphStyle('r', fontName='Helvetica',      fontSize=8,  leading=10, alignment=TA_RIGHT)

    oggi_fmt = date.today().strftime('%d/%m/%Y')
    story = []

    # ── 1. BANNER COPERTINA ──────────────────────────────────────
    banner = Table([[Paragraph(f'RENDICONTO ANNUALE {anno_str}', sTitle)]], colWidths=[W])
    banner.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), C_DARK),
        ('TOPPADDING',  (0,0), (-1,-1), 14),
        ('BOTTOMPADDING',(0,0),(-1,-1), 14),
        ('ROUNDEDCORNERS', [4]),
    ]))
    story.append(banner)
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(f'Generato il {oggi_fmt}', sSub))
    story.append(Spacer(1, 6*mm))

    # ── 2. RIEPILOGO ────────────────────────────────────────────
    sGreenBig = ParagraphStyle('gb', fontName='Helvetica-Bold', fontSize=14, textColor=C_GREEN, alignment=TA_CENTER)
    sRedBig   = ParagraphStyle('rb', fontName='Helvetica-Bold', fontSize=14, textColor=C_RED,   alignment=TA_CENTER)
    sDiffBig  = ParagraphStyle('db', fontName='Helvetica-Bold', fontSize=14,
                                textColor=(C_GREEN if riepilogo['differenza'] >= 0 else C_RED), alignment=TA_CENTER)
    sLabel    = ParagraphStyle('lb', fontName='Helvetica',      fontSize=7, textColor=C_MUTED,  alignment=TA_CENTER)
    W3 = W / 3
    riepilogo_table = Table([
        [Paragraph('TOTALE ENTRATE', sLabel), Paragraph('TOTALE USCITE', sLabel), Paragraph('DIFFERENZA', sLabel)],
        [Paragraph(_n(riepilogo['tot_entrate']), sGreenBig),
         Paragraph(_n(riepilogo['tot_uscite']),  sRedBig),
         Paragraph(('+' if riepilogo['differenza'] >= 0 else '') + _n(riepilogo['differenza']), sDiffBig)],
    ], colWidths=[W3, W3, W3])
    riepilogo_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), C_HEADER_ROW),
        ('BOX',           (0,0), (-1,-1), 0.5, C_MUTED),
        ('INNERGRID',     (0,0), (-1,-1), 0.5, C_MUTED),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(riepilogo_table)
    story.append(Spacer(1, 8*mm))

    # ── helper stili tabella ─────────────────────────────────────
    BASE_GRID = [('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
                 ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                 ('FONTSIZE', (0,0), (-1,-1), 7),
                 ('TOPPADDING', (0,0), (-1,-1), 3),
                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                 ('ALIGN', (1,0), (-1,-1), 'RIGHT')]

    # ── 3. TABELLA ENTRATE ───────────────────────────────────────
    story.append(Paragraph('Dettaglio Entrate', ParagraphStyle('h2', fontName='Helvetica-Bold', fontSize=10, leading=14, spaceBefore=4)))
    story.append(Spacer(1, 2*mm))

    # Larghezze colonne entrate: 60 + 14×12 + 25 + 24 = 277mm
    W_E = [60*mm] + [14*mm]*12 + [25*mm, 24*mm]
    hdr_e = ['Voce'] + MESI_BREVI + ['Tot.Pag.', 'Residuo']
    rows_e = [hdr_e]
    COLORE_SEZ = {'paghe': C_PAGHE, 'cont': C_CONT, 'paghe_cont': C_PAGHE_CONT, 'altro': C_ALTRO}
    style_e = list(BASE_GRID) + [
        ('BACKGROUND', (0,0), (-1,0), C_DARK),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 7),
    ]
    idx = 1

    for sez in entrate['sezioni_clienti']:
        bg = COLORE_SEZ.get(sez['colore'], C_ALTRO)
        row_sez = [sez['nome']] + [_n(v) for v in sez['subtotali_mesi']] + [_n(sez['subtotale']), _n(sez['subtotale_residui'])]
        rows_e.append(row_sez)
        style_e += [('BACKGROUND', (0,idx), (-1,idx), bg),
                    ('FONTNAME',   (0,idx), (-1,idx), 'Helvetica-Bold'),
                    ('TEXTCOLOR',  (0,idx), (-1,idx), C_DARK)]
        idx += 1
        for c in sez['clienti']:
            row_c = ['  › ' + c['nome']] + [_n(v) for v in c['mesi']] + [_n(c['tot_pagato'])]
            res_v = c['residuo']
            row_c.append(_n(res_v))
            rows_e.append(row_c)
            if res_v > 0:
                style_e.append(('TEXTCOLOR', (14, idx), (14, idx), C_RED))
            elif res_v < 0:
                style_e.append(('TEXTCOLOR', (14, idx), (14, idx), C_GREEN))
            idx += 1

    for mg in entrate['macrogruppi']:
        row_mg = [mg['nome']] + [_n(v) for v in mg['subtotali_mesi']] + [_n(mg['subtotale']), '—']
        rows_e.append(row_mg)
        style_e += [('BACKGROUND', (0,idx), (-1,idx), C_MACRO),
                    ('FONTNAME',   (0,idx), (-1,idx), 'Helvetica-Bold')]
        idx += 1
        for sv in mg['sottovoci']:
            rows_e.append(['  › ' + sv['nome']] + [_n(v) for v in sv['mesi']] + [_n(sv['totale']), '—'])
            idx += 1

    # Riga totale
    row_tot = ['TOTALE GENERALE'] + [_n(v) for v in entrate['totali_mesi']] + [_n(entrate['totale_annuale']), _n(entrate['totale_residui'])]
    rows_e.append(row_tot)
    style_e += [('BACKGROUND', (0,idx), (-1,idx), C_TOTAL),
                ('TEXTCOLOR',  (0,idx), (-1,idx), colors.white),
                ('FONTNAME',   (0,idx), (-1,idx), 'Helvetica-Bold')]

    t_e = Table(rows_e, colWidths=W_E, repeatRows=1)
    t_e.setStyle(TableStyle(style_e))
    story.append(t_e)
    story.append(Spacer(1, 8*mm))

    # ── 4. TABELLA USCITE ────────────────────────────────────────
    story.append(Paragraph('Dettaglio Uscite', ParagraphStyle('h2u', fontName='Helvetica-Bold', fontSize=10, leading=14, spaceBefore=4)))
    story.append(Spacer(1, 2*mm))

    # Larghezze: 80 + 14×12 + 29 = 277mm
    W_U = [80*mm] + [14*mm]*12 + [29*mm]
    rows_u = [['Voce'] + MESI_BREVI + ['Totale']]
    style_u = list(BASE_GRID) + [
        ('BACKGROUND', (0,0), (-1,0), C_DARK),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
    ]
    idx = 1
    for mg in uscite['macrogruppi']:
        rows_u.append([mg['nome']] + [_n(v) for v in mg['subtotali_mesi']] + [_n(mg['subtotale'])])
        style_u += [('BACKGROUND', (0,idx), (-1,idx), C_MACRO),
                    ('FONTNAME',   (0,idx), (-1,idx), 'Helvetica-Bold'),
                    ('TEXTCOLOR',  (1,idx), (-1,idx), C_RED)]
        idx += 1
        for sv in mg['sottovoci']:
            rows_u.append(['  › ' + sv['nome']] + [_n(v) for v in sv['mesi']] + [_n(sv['totale'])])
            style_u.append(('TEXTCOLOR', (1,idx), (-1,idx), C_RED))
            idx += 1
    rows_u.append(['TOTALE GENERALE'] + [_n(v) for v in uscite['totali_mesi']] + [_n(uscite['totale_annuale'])])
    style_u += [('BACKGROUND', (0,idx), (-1,idx), C_TOTAL),
                ('TEXTCOLOR',  (0,idx), (-1,idx), colors.white),
                ('FONTNAME',   (0,idx), (-1,idx), 'Helvetica-Bold'),
                ('TEXTCOLOR',  (1,idx), (-1,idx), colors.HexColor('#fca5a5'))]

    t_u = Table(rows_u, colWidths=W_U, repeatRows=1)
    t_u.setStyle(TableStyle(style_u))
    story.append(t_u)
    story.append(Spacer(1, 8*mm))

    # ── 5. TABELLA GIROCONTI ─────────────────────────────────────
    if giroconti:
        story.append(Paragraph('Giroconti', ParagraphStyle('h2g', fontName='Helvetica-Bold', fontSize=10, leading=14, spaceBefore=4)))
        story.append(Spacer(1, 2*mm))
        # Larghezze: 25+50+50+90+62 = 277mm
        W_G = [25*mm, 50*mm, 50*mm, 90*mm, 62*mm]
        rows_g = [['Data', 'Da', 'A', 'Descrizione', 'Importo']]
        for r in giroconti:
            rows_g.append([r['data'], r['da'], r['a'], r['descrizione'], _n(r['importo'])])
        style_g = list(BASE_GRID) + [
            ('BACKGROUND', (0,0), (-1,0), C_DARK),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('TEXTCOLOR',  (4,1), (4,-1), C_PURPLE),
            ('FONTNAME',   (4,1), (4,-1), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, C_HEADER_ROW]),
        ]
        t_g = Table(rows_g, colWidths=W_G, repeatRows=1)
        t_g.setStyle(TableStyle(style_g))
        story.append(t_g)

    # ── footer su ogni pagina ────────────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(C_MUTED)
        canvas.drawString(M, 6*mm, f'Documento generato il {oggi_fmt}')
        canvas.drawRightString(PAGE[0] - M, 6*mm, f'Pagina {doc.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
# GET /api/rendiconto/export-excel?anno=XXXX
# Genera Rendiconto_<anno>.xlsx  (4 fogli)
# ─────────────────────────────────────────────────────────────────
@bp.get('/api/rendiconto/export-excel')
@login_required
def export_excel():
    anno     = request.args.get('anno', str(date.today().year))
    anno_str = str(anno)
    db       = get_db()

    saldi_list = _query_saldi(db)
    riepilogo  = _query_riepilogo(db, anno_str)
    entrate    = _query_entrate(db, anno_str, mostra_archiviati=True)
    uscite     = _query_uscite(db, anno_str)
    giroconti  = _query_giroconti(db, anno_str)
    db.close()

    buf = _genera_excel(anno_str, saldi_list, riepilogo, entrate, uscite, giroconti)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'Rendiconto_{anno_str}.xlsx')


def _genera_excel(anno_str, saldi_list, riepilogo, entrate, uscite, giroconti):
    MESI = ['Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
            'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    FMT_EURO = '#,##0.00'
    oggi_fmt = date.today().strftime('%d/%m/%Y')

    # Stili riutilizzabili
    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _font(bold=False, color='000000', size=10):
        return Font(bold=bold, color=color, size=size)

    def _align(h='left', wrap=False):
        return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

    BORDER_THIN = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    F_DARK   = _fill('1A202C'); F_HDR = _fill('F1F5F9')
    F_PAGHE  = _fill('DCFCE7'); F_CONT = _fill('DBEAFE')
    F_PC     = _fill('FED7AA'); F_ALTRO = _fill('F1F5F9')
    F_MACRO  = _fill('E2E8F0'); F_TOTAL = _fill('1E293B')
    F_WHITE  = _fill('FFFFFF')

    def _style_row(ws, row_idx, col_start, col_end, fill=None, bold=False,
                   font_color='000000', num_fmt=None, align_right_from=None):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row_idx, column=c)
            if fill:  cell.fill = fill
            cell.font = _font(bold=bold, color=font_color)
            cell.border = BORDER_THIN
            if num_fmt and c >= (align_right_from or col_start):
                cell.number_format = num_fmt
                cell.alignment = _align('right')
            else:
                cell.alignment = _align('left')

    wb = openpyxl.Workbook()

    # ── FOGLIO 1: RIEPILOGO ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Riepilogo'
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18

    ws1.append([f'RENDICONTO ANNUALE {anno_str}'])
    ws1['A1'].font = _font(bold=True, size=14)
    ws1.append([f'Generato il {oggi_fmt}'])
    ws1['A2'].font = _font(color='64748B')
    ws1.append([])
    ws1.append(['Saldi conti'])
    ws1['A4'].font = _font(bold=True, size=11)

    saldo_start = 5
    for s in saldi_list:
        ws1.append([s['nome'], s['saldo']])
        r = ws1.max_row
        ws1[f'A{r}'].font = _font()
        ws1[f'B{r}'].number_format = FMT_EURO
        ws1[f'B{r}'].alignment = _align('right')

    saldo_end = ws1.max_row
    ws1.append(['Totale Disponibilità',
                f'=SUM(B{saldo_start}:B{saldo_end})'])
    r = ws1.max_row
    ws1[f'A{r}'].font = _font(bold=True)
    ws1[f'B{r}'].font = _font(bold=True)
    ws1[f'B{r}'].number_format = FMT_EURO
    ws1[f'B{r}'].alignment = _align('right')
    ws1[f'A{r}'].fill = F_HDR; ws1[f'B{r}'].fill = F_HDR

    ws1.append([])
    ws1.append(['Riepilogo annuale'])
    ws1[f'A{ws1.max_row}'].font = _font(bold=True, size=11)

    r_e = ws1.max_row + 1
    ws1.append(['Totale Entrate', riepilogo['tot_entrate']])
    ws1[f'B{r_e}'].number_format = FMT_EURO
    ws1[f'B{r_e}'].font = _font(bold=True, color='16A34A')
    ws1[f'B{r_e}'].alignment = _align('right')

    r_u = ws1.max_row + 1
    ws1.append(['Totale Uscite', riepilogo['tot_uscite']])
    ws1[f'B{r_u}'].number_format = FMT_EURO
    ws1[f'B{r_u}'].font = _font(bold=True, color='DC2626')
    ws1[f'B{r_u}'].alignment = _align('right')

    ws1.append(['Differenza', f'=B{r_e}-B{r_u}'])
    r = ws1.max_row
    ws1[f'A{r}'].font = _font(bold=True)
    ws1[f'B{r}'].number_format = FMT_EURO
    ws1[f'B{r}'].font = _font(bold=True)
    ws1[f'B{r}'].alignment = _align('right')
    ws1[f'A{r}'].fill = F_HDR; ws1[f'B{r}'].fill = F_HDR

    # ── FOGLIO 2: ENTRATE ────────────────────────────────────────
    ws2 = wb.create_sheet('Entrate')
    HDR_E = ['Voce', 'Tipo'] + MESI + ['Tot. Pagato', 'Residuo']
    ws2.append(HDR_E)
    # Header style
    for c in range(1, len(HDR_E) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = F_DARK
        cell.font = _font(bold=True, color='FFFFFF', size=9)
        cell.alignment = _align('center' if c == 1 else 'right')
        cell.border = BORDER_THIN

    SEZ_FILL = {'paghe': F_PAGHE, 'cont': F_CONT, 'paghe_cont': F_PC, 'altro': F_ALTRO}
    SEZ_TIPO = {'paghe': 'Paghe', 'cont': 'Contabilità', 'paghe_cont': 'Paghe+Cont.', 'altro': 'Altro'}

    def _append_money_row(ws, values, fill, bold, font_color='000000', indent=''):
        row_vals = [indent + str(values[0]), values[1]] + [v if v != 0 else None for v in values[2:]]
        ws.append(row_vals)
        ri = ws.max_row
        _style_row(ws, ri, 1, len(HDR_E), fill=fill, bold=bold,
                   font_color=font_color, num_fmt=FMT_EURO, align_right_from=3)
        ws.cell(ri, 1).alignment = _align('left')
        ws.cell(ri, 2).alignment = _align('center')
        return ri

    for sez in entrate['sezioni_clienti']:
        sfill = SEZ_FILL.get(sez['colore'], F_ALTRO)
        # riga subtotale sezione
        sub_row = [sez['nome'], SEZ_TIPO.get(sez['colore'], '')] + sez['subtotali_mesi'] + [sez['subtotale'], sez['subtotale_residui']]
        ri = _append_money_row(ws2, sub_row, sfill, True)
        # Formula subtotale (colonne 3..14 → mesi)
        ncol = len(HDR_E)
        for sez_c_idx in range(1, len(sez['clienti']) + 1):
            pass  # lasciamo i valori precalcolati

        for c in sez['clienti']:
            crow = ['  › ' + c['nome'], SEZ_TIPO.get(sez['colore'], '')] + c['mesi'] + [c['tot_pagato'], c['residuo']]
            ri = _append_money_row(ws2, crow, F_WHITE, False)
            # Formattazione condizionale residuo (col 16)
            res_cell = ws2.cell(ri, 16)
            if c['residuo'] > 0:
                res_cell.font = _font(color='DC2626')
            elif c['residuo'] < 0:
                res_cell.font = _font(color='16A34A')

    for mg in entrate['macrogruppi']:
        mg_row = [mg['nome'], ''] + mg['subtotali_mesi'] + [mg['subtotale'], None]
        _append_money_row(ws2, mg_row, F_MACRO, True)
        for sv in mg['sottovoci']:
            sv_row = ['  › ' + sv['nome'], ''] + sv['mesi'] + [sv['totale'], None]
            _append_money_row(ws2, sv_row, F_WHITE, False)

    # Riga TOTALE GENERALE
    last_data = ws2.max_row
    tot_row = ['TOTALE GENERALE', ''] + entrate['totali_mesi'] + [entrate['totale_annuale'], entrate['totale_residui']]
    ws2.append(tot_row)
    ri = ws2.max_row
    _style_row(ws2, ri, 1, len(HDR_E), fill=F_TOTAL, bold=True,
               font_color='FFFFFF', num_fmt=FMT_EURO, align_right_from=3)
    ws2.cell(ri, 1).alignment = _align('left')

    # Auto-filter, freeze, larghezze
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(HDR_E))}1'
    ws2.freeze_panes = 'C2'
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 14
    for i in range(3, 17):  ws2.column_dimensions[get_column_letter(i)].width = 12
    ws2.column_dimensions[get_column_letter(15)].width = 14
    ws2.column_dimensions[get_column_letter(16)].width = 14

    # ── FOGLIO 3: USCITE ─────────────────────────────────────────
    ws3 = wb.create_sheet('Uscite')
    HDR_U = ['Voce', 'Macrogruppo'] + MESI + ['Totale']
    ws3.append(HDR_U)
    for c in range(1, len(HDR_U) + 1):
        cell = ws3.cell(row=1, column=c)
        cell.fill = F_DARK; cell.font = _font(bold=True, color='FFFFFF', size=9)
        cell.alignment = _align('center' if c <= 2 else 'right')
        cell.border = BORDER_THIN

    for mg in uscite['macrogruppi']:
        mg_row = [mg['nome'], mg['nome']] + mg['subtotali_mesi'] + [mg['subtotale']]
        ws3.append(mg_row)
        ri = ws3.max_row
        _style_row(ws3, ri, 1, len(HDR_U), fill=F_MACRO, bold=True,
                   font_color='DC2626', num_fmt=FMT_EURO, align_right_from=3)
        ws3.cell(ri, 1).font = _font(bold=True, color='000000')
        ws3.cell(ri, 1).alignment = _align('left')
        ws3.cell(ri, 2).alignment = _align('left')
        for sv in mg['sottovoci']:
            sv_row = ['  › ' + sv['nome'], mg['nome']] + sv['mesi'] + [sv['totale']]
            ws3.append(sv_row)
            ri = ws3.max_row
            _style_row(ws3, ri, 1, len(HDR_U), fill=F_WHITE, bold=False,
                       font_color='DC2626', num_fmt=FMT_EURO, align_right_from=3)
            ws3.cell(ri, 1).font = _font(color='000000')
            ws3.cell(ri, 1).alignment = _align('left')
            ws3.cell(ri, 2).alignment = _align('left')

    ws3.append(['TOTALE GENERALE', ''] + uscite['totali_mesi'] + [uscite['totale_annuale']])
    ri = ws3.max_row
    _style_row(ws3, ri, 1, len(HDR_U), fill=F_TOTAL, bold=True,
               font_color='FFFFFF', num_fmt=FMT_EURO, align_right_from=3)
    ws3.cell(ri, 1).alignment = _align('left')
    ws3.cell(ri, len(HDR_U)).font = _font(bold=True, color='FCA5A5')

    ws3.auto_filter.ref = f'A1:{get_column_letter(len(HDR_U))}1'
    ws3.freeze_panes = 'C2'
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 18
    for i in range(3, len(HDR_U) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 12

    # ── FOGLIO 4: GIROCONTI ──────────────────────────────────────
    ws4 = wb.create_sheet('Giroconti')
    HDR_G = ['Data', 'Da', 'A', 'Descrizione', 'Importo']
    ws4.append(HDR_G)
    for c in range(1, 6):
        cell = ws4.cell(row=1, column=c)
        cell.fill = F_DARK; cell.font = _font(bold=True, color='FFFFFF', size=9)
        cell.alignment = _align('center' if c < 5 else 'right')
        cell.border = BORDER_THIN

    alt = False
    for g in giroconti:
        ws4.append([g['data'], g['da'], g['a'], g['descrizione'], g['importo']])
        ri = ws4.max_row
        fill = _fill('F8FAFC') if alt else F_WHITE
        for c in range(1, 6):
            cell = ws4.cell(ri, c)
            cell.fill = fill; cell.border = BORDER_THIN
            cell.alignment = _align('right' if c == 5 else 'left')
        ws4.cell(ri, 5).number_format = FMT_EURO
        ws4.cell(ri, 5).font = _font(color='7C3AED', bold=True)
        alt = not alt

    ws4.auto_filter.ref = 'A1:E1'
    ws4.freeze_panes = 'A2'
    widths = [12, 20, 20, 40, 16]
    for i, w in enumerate(widths, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
