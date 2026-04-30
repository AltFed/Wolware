# routes/strumenti.py
from flask import Blueprint, request, jsonify, send_file
from auth.routes import login_required
from database import get_db
import json
import io

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

bp = Blueprint('strumenti_bp', __name__)

# ─────────────────────────────────────────────
# TIPI MACROGRUPPO
# ─────────────────────────────────────────────
TIPI_FISSI     = ('costi_fissi_mensili', 'costi_fissi_annuali')
TIPI_VARIABILI = ('costi_variabili_mensili', 'costi_variabili_annuali')
MESI_NOMI = ['', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio',
             'Giugno', 'Luglio', 'Agosto', 'Settembre', 'Ottobre',
             'Novembre', 'Dicembre']


# ═══════════════════════════════════════════════════════════
#  BLOCCO 4.4 — CONTABILIZZA COSTI FISSI
# ═══════════════════════════════════════════════════════════

def _voce_attiva_nel_mese(voce, mese: int, tipo: str) -> bool:
    """
    Costi Fissi Mensili → attivi in tutti i 12 mesi.
    Costi Fissi Annuali → attivi solo nei mesi presenti in mesi_json.
    """
    if tipo == 'costi_fissi_mensili':
        return True
    # annuali: controlla mesi_json
    mesi_json = voce['mesi_json']
    if not mesi_json:
        return False
    try:
        mesi = json.loads(mesi_json) if isinstance(mesi_json, str) else mesi_json
        return mese in mesi
    except Exception:
        return False


def _gestione_attiva(ditta, anno: int, tipo_gestione: str) -> bool:
    """Verifica se la gestione (paghe/contabilita) era attiva nell'anno."""
    if tipo_gestione == 'paghe':
        inizio = ditta.get('inizio_paghe') or ditta.get('inizioPaghe')
        fine   = ditta.get('fine_paghe')   or ditta.get('finePaghe')
    else:
        inizio = ditta.get('inizio_contabilita') or ditta.get('inizioContabilita')
        fine   = ditta.get('fine_contabilita')   or ditta.get('fineContabilita')

    if not inizio:
        return False
    # Semplice: basta che l'anno sia ≥ anno inizio e ≤ anno fine (o fine assente)
    try:
        anno_inizio = int(str(inizio)[:4])
    except Exception:
        return True
    if anno < anno_inizio:
        return False
    if fine:
        try:
            anno_fine = int(str(fine)[:4])
            if anno > anno_fine:
                return False
        except Exception:
            pass
    return True


def _gia_contabilizzata(db, ditta_id: int, voce_id: int, anno: int, mese: int) -> bool:
    row = db.execute(
        '''SELECT id FROM pratiche
           WHERE ditta_id=? AND voce_costo_id=? AND anno=? AND mese=?
             AND tipo IN ('costi_fissi_mensili','costi_fissi_annuali')''',
        (ditta_id, voce_id, anno, mese)
    ).fetchone()
    return row is not None


@bp.route('/api/strumenti/contabilizza/preview', methods=['GET'])
@login_required
def contabilizza_preview():
    """
    Restituisce l'anteprima della contabilizzazione.
    Query params: anno, mese_da, mese_a, ditta_id (opzionale)
    """
    anno    = int(request.args.get('anno', 2026))
    mese_da = int(request.args.get('mese_da', 1))
    mese_a  = int(request.args.get('mese_a', 12))
    ditta_id_filter = request.args.get('ditta_id')

    db = get_db()
    try:
        if ditta_id_filter:
            ditte = db.execute(
                'SELECT * FROM ditte WHERE id=? AND archiviato=0',
                (int(ditta_id_filter),)
            ).fetchall()
        else:
            ditte = db.execute(
                'SELECT * FROM ditte WHERE archiviato=0 ORDER BY denominazione COLLATE NOCASE'
            ).fetchall()

        righe = []
        tot_da_fare = 0
        tot_gia_ok  = 0
        tot_importo = 0.0

        for ditta in ditte:
            voci = db.execute(
                '''SELECT dv.*, mg.tipo as mg_tipo
                   FROM ditta_voci dv
                   JOIN macrogruppi mg ON mg.id = dv.macrogruppo_id
                   WHERE dv.ditta_id=? AND mg.tipo IN (?,?)''',
                (ditta['id'], 'costi_fissi_mensili', 'costi_fissi_annuali')
            ).fetchall()

            for voce in voci:
                tipo = voce['mg_tipo']
                # flag anno precedente
                richiede_anno_prec = voce['richiede_anno_precedente'] if 'richiede_anno_precedente' in voce.keys() else 0
                if richiede_anno_prec:
                    if not _gestione_attiva(dict(ditta), anno - 1, 'paghe'):
                        continue

                for mese in range(mese_da, mese_a + 1):
                    if not _voce_attiva_nel_mese(dict(voce), mese, tipo):
                        continue

                    gia_ok = _gia_contabilizzata(db, ditta['id'], voce['voce_costo_id'] or voce['id'], anno, mese)
                    stato  = 'ok' if gia_ok else 'da_fare'
                    importo = float(voce['prezzo'] or 0)

                    righe.append({
                        'ditta_id':   ditta['id'],
                        'ditta_nome': ditta['denominazione'],
                        'voce_id':    voce['voce_costo_id'] or voce['id'],
                        'voce_nome':  voce['nome'],
                        'macrogruppo_nome': voce['macrogruppo_nome'],
                        'tipo':       tipo,
                        'mese':       mese,
                        'mese_nome':  MESI_NOMI[mese],
                        'anno':       anno,
                        'qta':        1,
                        'prezzo':     importo,
                        'totale':     importo,
                        'esente_iva': bool(voce['esente_iva']),
                        'stato':      stato,
                    })

                    if gia_ok:
                        tot_gia_ok += 1
                    else:
                        tot_da_fare += 1
                        tot_importo += importo

        return jsonify({
            'righe':       righe,
            'tot_da_fare': tot_da_fare,
            'tot_gia_ok':  tot_gia_ok,
            'tot_importo': round(tot_importo, 2),
        })
    finally:
        db.close()


@bp.route('/api/strumenti/contabilizza', methods=['POST'])
@login_required
def contabilizza_esegui():
    """
    Esegue la contabilizzazione delle voci 'da_fare'.
    Idempotente: salta quelle già presenti.
    Body: { anno, mese_da, mese_a, ditta_id? }
    """
    d = request.get_json()
    anno        = int(d.get('anno', 2026))
    mese_da     = int(d.get('mese_da', 1))
    mese_a      = int(d.get('mese_a', 12))
    ditta_filter = d.get('ditta_id')

    db = get_db()
    try:
        if ditta_filter:
            ditte = db.execute(
                'SELECT * FROM ditte WHERE id=? AND archiviato=0',
                (int(ditta_filter),)
            ).fetchall()
        else:
            ditte = db.execute('SELECT * FROM ditte WHERE archiviato=0').fetchall()

        aggiunte = saltate = 0

        for ditta in ditte:
            voci = db.execute(
                '''SELECT dv.*, mg.tipo as mg_tipo
                   FROM ditta_voci dv
                   JOIN macrogruppi mg ON mg.id = dv.macrogruppo_id
                   WHERE dv.ditta_id=? AND mg.tipo IN (?,?)''',
                (ditta['id'], 'costi_fissi_mensili', 'costi_fissi_annuali')
            ).fetchall()

            for voce in voci:
                tipo = voce['mg_tipo']
                richiede_anno_prec = voce['richiede_anno_precedente'] if 'richiede_anno_precedente' in voce.keys() else 0
                if richiede_anno_prec:
                    if not _gestione_attiva(dict(ditta), anno - 1, 'paghe'):
                        continue

                for mese in range(mese_da, mese_a + 1):
                    if not _voce_attiva_nel_mese(dict(voce), mese, tipo):
                        continue

                    ref_voce_id = voce['voce_costo_id'] or voce['id']
                    if _gia_contabilizzata(db, ditta['id'], ref_voce_id, anno, mese):
                        saltate += 1
                        continue

                    db.execute(
                        '''INSERT INTO pratiche
                           (ditta_id, voce_costo_id, anno, mese, descrizione,
                            qta, prezzo, importo, tipo,
                            macrogruppo_nome, esente_iva, data_inserimento)
                           VALUES (?,?,?,?,?,1,?,?,?,?,?,date('now'))''',
                        (
                            ditta['id'],
                            ref_voce_id,
                            anno, mese,
                            voce['nome'],
                            float(voce['prezzo'] or 0),
                            float(voce['prezzo'] or 0),
                            tipo,
                            voce['macrogruppo_nome'],
                            int(voce['esente_iva'] or 0),
                        )
                    )
                    aggiunte += 1

        db.commit()
        return jsonify({'ok': True, 'aggiunte': aggiunte, 'saltate': saltate})
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════
#  BLOCCO 4.3 — INSERIMENTO COSTI VARIABILI
# ═══════════════════════════════════════════════════════════

@bp.route('/api/strumenti/variabili/tabella', methods=['GET'])
@login_required
def variabili_tabella():
    """
    Restituisce la struttura tabella:
    - colonne: tutte le voci variabili attive nel mese/anno
    - righe: un record per ditta con le celle compilabili
    Query params: anno, mese
    """
    anno = int(request.args.get('anno', 2026))
    mese = int(request.args.get('mese', 1))

    db = get_db()
    try:
        # Raccoglie tutte le voci variabili distinte
        tutte_voci = db.execute(
            '''SELECT DISTINCT dv.voce_costo_id, dv.nome, dv.prezzo,
                      dv.macrogruppo_nome, mg.tipo as mg_tipo, dv.mesi_json
               FROM ditta_voci dv
               JOIN macrogruppi mg ON mg.id = dv.macrogruppo_id
               JOIN ditte d ON d.id = dv.ditta_id
               WHERE mg.tipo IN (?,?) AND d.archiviato=0''',
            ('costi_variabili_mensili', 'costi_variabili_annuali')
        ).fetchall()

        # Filtra per mese: mensili sempre visibili, annuali solo nei mesi configurati
        colonne = []
        colonne_ids = set()
        for v in tutte_voci:
            if v['voce_costo_id'] in colonne_ids:
                continue
            attiva = _voce_attiva_nel_mese(dict(v), mese, v['mg_tipo'])
            if attiva:
                colonne.append({
                    'voce_id':   v['voce_costo_id'],
                    'nome':      v['nome'],
                    'mg_nome':   v['macrogruppo_nome'],
                    'tipo':      v['mg_tipo'],
                })
                colonne_ids.add(v['voce_costo_id'])

        # Per ogni ditta costruisce il record riga
        ditte = db.execute(
            'SELECT * FROM ditte WHERE archiviato=0 ORDER BY denominazione COLLATE NOCASE'
        ).fetchall()

        righe = []
        for ditta in ditte:
            celle = {}
            for col in colonne:
                # Verifica che la ditta abbia questa voce
                dv = db.execute(
                    '''SELECT dv.id, dv.prezzo, dv.mesi_json
                       FROM ditta_voci dv
                       JOIN macrogruppi mg ON mg.id = dv.macrogruppo_id
                       WHERE dv.ditta_id=? AND dv.voce_costo_id=?
                         AND mg.tipo IN (?,?)''',
                    (ditta['id'], col['voce_id'],
                     'costi_variabili_mensili', 'costi_variabili_annuali')
                ).fetchone()

                if dv and _voce_attiva_nel_mese(dict(dv), mese, col['tipo']):
                    # Cerca quantità già inserita per questo mese
                    esistente = db.execute(
                        '''SELECT qta FROM pratiche
                           WHERE ditta_id=? AND voce_costo_id=? AND anno=? AND mese=?
                             AND tipo IN (?,?)''',
                        (ditta['id'], col['voce_id'], anno, mese,
                         'costi_variabili_mensili', 'costi_variabili_annuali')
                    ).fetchone()
                    celle[col['voce_id']] = {
                        'attiva':  True,
                        'qta':     esistente['qta'] if esistente else 0,
                        'prezzo':  float(dv['prezzo'] or 0),
                    }
                else:
                    celle[col['voce_id']] = {'attiva': False, 'qta': None, 'prezzo': None}

            righe.append({
                'ditta_id':   ditta['id'],
                'ditta_nome': ditta['denominazione'],
                'celle':      celle,
            })

        return jsonify({'colonne': colonne, 'righe': righe})
    finally:
        db.close()


@bp.route('/api/strumenti/variabili/carica', methods=['POST'])
@login_required
def variabili_carica():
    """
    Salva le quantità costi variabili nel mese.
    Body: { anno, mese, dati: [ { ditta_id, voce_id, qta } ] }
    Comportamento: upsert (aggiorna se esiste, inserisce se no).
    Se qta == 0: rimuove la riga (nessun costo).
    """
    d = request.get_json()
    anno = int(d.get('anno', 2026))
    mese = int(d.get('mese', 1))
    dati = d.get('dati', [])

    db = get_db()
    try:
        salvati = eliminati = 0

        for item in dati:
            ditta_id = int(item['ditta_id'])
            voce_id  = int(item['voce_id'])
            qta      = int(item.get('qta', 0))

            esistente = db.execute(
                '''SELECT id FROM pratiche
                   WHERE ditta_id=? AND voce_costo_id=? AND anno=? AND mese=?
                     AND tipo IN (?,?)''',
                (ditta_id, voce_id, anno, mese,
                 'costi_variabili_mensili', 'costi_variabili_annuali')
            ).fetchone()

            if qta == 0:
                if esistente:
                    db.execute('DELETE FROM pratiche WHERE id=?', (esistente['id'],))
                    eliminati += 1
                continue

            # Recupera info voce
            voce_info = db.execute(
                '''SELECT dv.nome, dv.prezzo, dv.macrogruppo_nome, dv.esente_iva,
                          mg.tipo as mg_tipo
                   FROM ditta_voci dv
                   JOIN macrogruppi mg ON mg.id = dv.macrogruppo_id
                   WHERE dv.ditta_id=? AND dv.voce_costo_id=?''',
                (ditta_id, voce_id)
            ).fetchone()
            if not voce_info:
                continue

            importo = round(float(voce_info['prezzo'] or 0) * qta, 2)

            if esistente:
                db.execute(
                    'UPDATE pratiche SET qta=?, importo=? WHERE id=?',
                    (qta, importo, esistente['id'])
                )
            else:
                db.execute(
                    '''INSERT INTO pratiche
                       (ditta_id, voce_costo_id, anno, mese, descrizione,
                        qta, prezzo, importo, tipo, macrogruppo_nome,
                        esente_iva, data_inserimento)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,date('now'))''',
                    (ditta_id, voce_id, anno, mese,
                     voce_info['nome'], qta,
                     float(voce_info['prezzo'] or 0), importo,
                     voce_info['mg_tipo'], voce_info['macrogruppo_nome'],
                     int(voce_info['esente_iva'] or 0))
                )
            salvati += 1

        db.commit()
        return jsonify({'ok': True, 'salvati': salvati, 'eliminati': eliminati})
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════
#  BLOCCO 4.2 — COSTI MASSIVI
# ═══════════════════════════════════════════════════════════

@bp.route('/api/strumenti/costi-massivi', methods=['POST'])
@login_required
def costi_massivi():
    """
    Inserisce una pratica a richiesta su più clienti.
    Body: {
      descrizione, importo, qta, mese, anno, esente_iva,
      selezione: 'tutti' | 'tariffario' | 'manuale',
      tariffario_id?,          # se selezione='tariffario'
      ditta_ids?               # se selezione='manuale' → lista di int
    }
    """
    d = request.get_json()
    descrizione = d.get('descrizione', '').strip()
    if not descrizione:
        return jsonify({'error': 'Descrizione obbligatoria'}), 400

    importo_unit = float(d.get('importo', 0))
    qta          = int(d.get('qta', 1))
    mese         = int(d.get('mese', 1))
    anno         = int(d.get('anno', 2026))
    esente_iva   = int(d.get('esente_iva', 0))
    selezione    = d.get('selezione', 'tutti')
    importo_tot  = round(importo_unit * qta, 2)

    db = get_db()
    try:
        if selezione == 'tutti':
            ditte = db.execute('SELECT id FROM ditte WHERE archiviato=0').fetchall()
        elif selezione == 'tariffario':
            tar_id = int(d.get('tariffario_id', 0))
            ditte  = db.execute(
                'SELECT id FROM ditte WHERE tariffario_id=? AND archiviato=0',
                (tar_id,)
            ).fetchall()
        else:  # manuale
            ids   = [int(i) for i in d.get('ditta_ids', [])]
            placeholder = ','.join('?' * len(ids))
            ditte = db.execute(
                f'SELECT id FROM ditte WHERE id IN ({placeholder}) AND archiviato=0',
                ids
            ).fetchall() if ids else []

        inseriti = 0
        for ditta in ditte:
            db.execute(
                '''INSERT INTO pratiche
                   (ditta_id, anno, mese, descrizione,
                    qta, prezzo, importo, tipo,
                    esente_iva, data_inserimento)
                   VALUES (?,?,?,?,?,?,?,'a_richiesta',?,date('now'))''',
                (ditta['id'], anno, mese, descrizione,
                 qta, importo_unit, importo_tot, esente_iva)
            )
            inseriti += 1

        db.commit()
        return jsonify({'ok': True, 'inseriti': inseriti, 'totale_euro': round(importo_tot * inseriti, 2)})
    finally:
        db.close()


@bp.route('/api/strumenti/costi-massivi/anteprima', methods=['POST'])
@login_required
def costi_massivi_anteprima():
    """Conta i clienti selezionati e calcola il totale. Nessuna scrittura."""
    d         = request.get_json()
    importo   = float(d.get('importo', 0))
    qta       = int(d.get('qta', 1))
    selezione = d.get('selezione', 'tutti')

    db = get_db()
    try:
        if selezione == 'tutti':
            n = db.execute('SELECT COUNT(*) FROM ditte WHERE archiviato=0').fetchone()[0]
        elif selezione == 'tariffario':
            tar_id = int(d.get('tariffario_id', 0))
            n = db.execute(
                'SELECT COUNT(*) FROM ditte WHERE tariffario_id=? AND archiviato=0',
                (tar_id,)
            ).fetchone()[0]
        else:
            ids = [int(i) for i in d.get('ditta_ids', [])]
            n   = len(ids)

        totale = round(importo * qta * n, 2)
        return jsonify({'clienti': n, 'totale': totale})
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════
#  BLOCCO 4.1 — IMPORTA CLIENTI DA EXCEL
# ═══════════════════════════════════════════════════════════

@bp.route('/api/strumenti/import/template', methods=['GET'])
@login_required
def import_template():
    """Genera e scarica il template Excel per l'import clienti."""
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl non installato'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clienti'

    headers = [
        'denominazione', 'codice_fiscale', 'email', 'telefono',
        'indirizzo', 'residuo_iniziale', 'cadenza_pagamenti',
        'inizio_paghe', 'fine_paghe', 'inizio_contabilita',
        'fine_contabilita', 'annotazioni',
    ]
    ws.append(headers)
    # Riga esempio
    ws.append([
        'Rossi S.r.l.', '12345678901', 'info@rossi.it', '0123456789',
        'Via Roma 1, 00100 Roma', 0, 'mensile',
        '2023-01-01', '', '2023-01-01', '', 'Cliente esempio',
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='template_import_clienti.xlsx'
    )


@bp.route('/api/strumenti/import/preview', methods=['POST'])
@login_required
def import_preview():
    """
    Carica file Excel e restituisce anteprima (senza salvare).
    Form: file=<xlsx>, tariffario_id=<int|''>
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl non installato'}), 500

    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400

    f = request.files['file']
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active

    CADENZE_VALIDE = {'mensile', 'trimestrale', 'quadrimestrale', 'semestrale', 'libero'}

    db = get_db()
    try:
        denom_esistenti = {
            r['denominazione'].lower()
            for r in db.execute('SELECT denominazione FROM ditte').fetchall()
        }

        headers = None
        pronti  = []
        errori  = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip().lower() if c else '' for c in row]
                continue

            record = dict(zip(headers, row))
            riga_n = i + 1  # 1-indexed per l'utente

            denom = (record.get('denominazione') or '').strip()
            if not denom:
                errori.append(f'Riga {riga_n}: denominazione mancante')
                continue
            if denom.lower() in denom_esistenti:
                errori.append(f'Riga {riga_n}: "{denom}" esiste già')
                continue

            cadenza = (record.get('cadenza_pagamenti') or 'libero').strip().lower()
            if cadenza not in CADENZE_VALIDE:
                cadenza = 'libero'

            pronti.append({
                'denominazione':    denom,
                'codice_fiscale':   str(record.get('codice_fiscale') or '').strip(),
                'email':            str(record.get('email') or '').strip(),
                'telefono':         str(record.get('telefono') or '').strip(),
                'indirizzo':        str(record.get('indirizzo') or '').strip(),
                'residuo_iniziale': float(record.get('residuo_iniziale') or 0),
                'cadenza_pagamenti': cadenza,
                'inizio_paghe':     str(record.get('inizio_paghe') or '').strip()[:10] or None,
                'fine_paghe':       str(record.get('fine_paghe') or '').strip()[:10] or None,
                'inizio_contabilita': str(record.get('inizio_contabilita') or '').strip()[:10] or None,
                'fine_contabilita':  str(record.get('fine_contabilita') or '').strip()[:10] or None,
                'annotazioni':      str(record.get('annotazioni') or '').strip(),
            })

        return jsonify({'pronti': pronti, 'errori': errori})
    finally:
        db.close()
        wb.close()


@bp.route('/api/strumenti/import/esegui', methods=['POST'])
@login_required
def import_esegui():
    """
    Salva i clienti dall'anteprima.
    Body: { clienti: [...], tariffario_id: int|null }
    """
    d            = request.get_json()
    clienti      = d.get('clienti', [])
    tariffario_id = d.get('tariffario_id') or None

    db = get_db()
    try:
        # Se c'è un tariffario, carica le voci da copiare
        voci_tar = []
        tar_nome  = None
        if tariffario_id:
            tar_row  = db.execute('SELECT nome FROM tariffari WHERE id=?', (tariffario_id,)).fetchone()
            tar_nome = tar_row['nome'] if tar_row else None
            voci_tar = db.execute(
                '''SELECT vc.*, mg.id as mg_id, mg.nome as mg_nome, mg.tipo
                   FROM voci_costo vc
                   JOIN macrogruppi mg ON mg.id = vc.macrogruppo_id
                   WHERE mg.tariffario_id=?''',
                (tariffario_id,)
            ).fetchall()

        inseriti = 0
        for c in clienti:
            db.execute(
                '''INSERT INTO ditte
                   (denominazione, codice_fiscale, email, telefono,
                    indirizzo, residuo_iniziale, cadenza_pagamenti,
                    inizio_paghe, fine_paghe,
                    inizio_contabilita, fine_contabilita,
                    annotazioni, tariffario_id, tariffario_nome,
                    archiviato)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)''',
                (
                    c['denominazione'], c.get('codice_fiscale'), c.get('email'),
                    c.get('telefono'), c.get('indirizzo'),
                    float(c.get('residuo_iniziale', 0)),
                    c.get('cadenza_pagamenti', 'libero'),
                    c.get('inizio_paghe') or None,
                    c.get('fine_paghe') or None,
                    c.get('inizio_contabilita') or None,
                    c.get('fine_contabilita') or None,
                    c.get('annotazioni', ''),
                    tariffario_id, tar_nome,
                )
            )
            nuova_ditta_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

            # Copia profonda voci tariffario
            TIPO_MAP = {
                'costi_fissi_mensili':    'costi_fissi_mensili',
                'costi_fissi_annuali':    'costi_fissi_annuali',
                'costi_variabili_mensili': 'costi_variabili_mensili',
                'costi_variabili_annuali': 'costi_variabili_annuali',
            }
            for v in voci_tar:
                db.execute(
                    '''INSERT INTO ditta_voci
                       (ditta_id, voce_costo_id, nome, prezzo, note,
                        macrogruppo_nome, macrogruppo_id, tipo, custom,
                        esente_iva, richiede_anno_precedente, mesi_json, sync_override)
                       VALUES (?,?,?,?,?,?,?,?,0,?,?,?,0)''',
                    (nuova_ditta_id, v['id'], v['nome'], v['prezzo'],
                     v.get('note', ''), v['mg_nome'], v['mg_id'],
                     TIPO_MAP.get(v['tipo'], v['tipo']),
                     v.get('esente_iva', 0), v.get('richiede_anno_precedente', 0),
                     v.get('mesi_json'))
                )
            inseriti += 1

        db.commit()
        return jsonify({'ok': True, 'inseriti': inseriti})
    finally:
        db.close()